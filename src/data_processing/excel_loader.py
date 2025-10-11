"""
Excel Data Loader Module

This module provides robust Excel file loading capabilities for Taiwan customs
import/export statistics. Handles irregular structures including Chinese headers,
merged cells, and multi-level column headers.

Author: Claude AI & 潘驄杰
Date: 2025-10-11
"""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, Tuple, Optional, List
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelLoader:
    """
    Excel file loader for Taiwan customs trade statistics.

    Handles:
    - Chinese headers and merged cells
    - Multi-level column headers
    - Automatic data row detection
    - Metadata extraction
    """

    def __init__(self, data_dir: Optional[Path] = None):
        """
        Initialize Excel loader.

        Args:
            data_dir: Path to data directory containing Excel files.
                     Defaults to data/August2025_PreliminaryStatistics_on_CustomsImports_and_Exports/
        """
        if data_dir is None:
            self.data_dir = Path("data/August2025_PreliminaryStatistics_on_CustomsImports_and_Exports")
        else:
            self.data_dir = Path(data_dir)

        if not self.data_dir.exists():
            raise FileNotFoundError(f"Data directory not found: {self.data_dir}")

    def detect_header_and_data_rows(self, file_path: Path, sheet_name: int = 0) -> Tuple[int, int]:
        """
        Detect the header row and data start row.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet index (default 0 for first sheet)

        Returns:
            Tuple of (header_row_index, data_start_row_index) - both 0-indexed
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if isinstance(sheet_name, int):
            sheet = wb.worksheets[sheet_name]
        else:
            sheet = wb[sheet_name]

        header_row = None
        data_row = None

        # Strategy:
        # 1. Find header row (contains "年(月)別" or similar header text)
        # 2. Find first data row (contains year pattern like "104年")
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True)):
            if row[0] is not None:
                cell_value = str(row[0]).strip()

                # Check for header row (contains header text patterns)
                if header_row is None and ("年(月)別" in cell_value or "年月別" in cell_value or cell_value == "年月"):
                    header_row = i
                    logger.debug(f"Header row found at row {i+1} (1-indexed)")

                # Check for data row (year pattern like "104年", "105年")
                if data_row is None and "年" in cell_value:
                    # Check if it's a year value (not a header like "年(月)別")
                    if cell_value[0:3].isdigit() or cell_value[0:2].isdigit():
                        data_row = i
                        logger.debug(f"Data starts at row {i+1} (1-indexed)")
                        break  # Found data row, stop searching

        # Fallback: use common patterns
        if header_row is None:
            header_row = 3  # 0-indexed (row 4 in Excel)
            logger.warning(f"Could not detect header row, using default row 4")

        if data_row is None:
            data_row = 5  # 0-indexed (row 6 in Excel)
            logger.warning(f"Could not detect data row, using default row 6")

        return header_row, data_row

    def extract_metadata(self, file_path: Path, sheet_name: int = 0) -> Dict[str, str]:
        """
        Extract metadata from Excel file (title, unit, etc.).

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet index

        Returns:
            Dictionary containing metadata
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if isinstance(sheet_name, int):
            sheet = wb.worksheets[sheet_name]
        else:
            sheet = wb[sheet_name]

        metadata = {
            "file_path": str(file_path),
            "file_name": file_path.name,
            "sheet_name": sheet.title,
        }

        # Extract title (usually in first row, first cell)
        title_cell = sheet.cell(row=1, column=1).value
        if title_cell:
            metadata["title"] = str(title_cell).strip()

        # Extract unit information (usually in row 3, somewhere in the row)
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell and "單位" in str(cell):
                    metadata["unit"] = str(cell).strip()
                    break

        return metadata

    def load_excel_table(
        self,
        table_id: str,
        skip_rows: Optional[int] = None,
        header_row: Optional[int] = None
    ) -> Tuple[pd.DataFrame, Dict[str, str]]:
        """
        Load Excel table with automatic structure detection.

        Args:
            table_id: Table identifier (e.g., "table01", "table08")
            skip_rows: Number of rows to skip (auto-detected if None)
            header_row: Row index for headers (auto-detected if None)

        Returns:
            Tuple of (DataFrame with raw data, metadata dictionary)

        Raises:
            FileNotFoundError: If table file doesn't exist
            ValueError: If table cannot be loaded
        """
        # Map table_id to file name
        table_file_map = {
            "table01": "Table1_Import_and_ExportTradeValues.xlsx",
            "table02": "Table2_Classification_of_MajorExportCommodities.xlsx",
            "table03": "Table3_ Classification_of_MajorImportedGoods.xlsx",
            "table04": "Table4_MajorExportCommodities.xlsx",
            "table05": "Table5_MajorImportedCommodities.xlsx",
            "table06": "Table6_ExportTradeStructure.xlsx",
            "table07": "Table7_ImportTradeStructure.xlsx",
            "table08": "Table8_Taiwans's_ExportValue_and_AnnualGrowthRate.xlsx",
            "table09": "Table9_ ImportValue_AnnualGrowthRate(to_Taiwan).xlsx",
            "table10": "Table10_Surplus_inTrade_with_MajorCountries.xlsx",
            "table11": "Table11_MajorExportCommodities(China_and_HongKong).xlsx",
            "table12": "Table12_ExporValue_and_AnnualGrowthRate_to_18Countries_UnderNewSouthbound_Policy.xlsx",
            "table13": "Table13_Seasonally_AdjustedImport_and_ExportTradeValues.xlsx",
            "table14": "Table14_Import_and_ExportValues_and_AnnualGrowthRates_for_MajorCountries__OR_Regions.xlsx",
            "table15": "Table15_Import_and_Export_Price-RelatedIndicators.xlsx",
            "table16": "Table16_ExchangeRates_of_MajorCountries_CurrenciesAgainst_USDollar.xlsx",
        }

        if table_id not in table_file_map:
            raise ValueError(f"Unknown table_id: {table_id}. Valid IDs: {list(table_file_map.keys())}")

        file_path = self.data_dir / table_file_map[table_id]

        if not file_path.exists():
            raise FileNotFoundError(f"Table file not found: {file_path}")

        logger.info(f"Loading {table_id} from {file_path.name}")

        # Extract metadata
        metadata = self.extract_metadata(file_path)
        metadata["table_id"] = table_id

        # Detect header and data rows
        header_row, data_row = self.detect_header_and_data_rows(file_path)

        # Load data with pandas
        # Skip rows before header, use header row, data starts after header
        try:
            # Skip all rows before the header row
            if header_row > 0:
                skip_list = list(range(0, header_row))
            else:
                skip_list = None

            df = pd.read_excel(
                file_path,
                sheet_name=0,
                skiprows=skip_list,  # Skip rows before header
                header=0,  # First non-skipped row is header
            )

            # Remove empty columns (all NaN)
            df = df.dropna(axis=1, how='all')

            # Remove empty rows (all NaN except first column)
            df = df.dropna(axis=0, how='all', subset=df.columns[1:])

            # Clean column names (remove whitespace, handle unnamed)
            df.columns = [str(col).strip() if not str(col).startswith('Unnamed') else f'col_{i}'
                         for i, col in enumerate(df.columns)]

            logger.info(f"Loaded {table_id}: shape={df.shape}, columns={len(df.columns)}")

            return df, metadata

        except Exception as e:
            logger.error(f"Failed to load {table_id}: {e}")
            raise ValueError(f"Failed to load {table_id}: {e}")

    def load_multiple_tables(
        self,
        table_ids: List[str]
    ) -> Dict[str, Tuple[pd.DataFrame, Dict[str, str]]]:
        """
        Load multiple Excel tables.

        Args:
            table_ids: List of table identifiers

        Returns:
            Dictionary mapping table_id to (DataFrame, metadata) tuples
        """
        results = {}
        for table_id in table_ids:
            try:
                df, metadata = self.load_excel_table(table_id)
                results[table_id] = (df, metadata)
            except Exception as e:
                logger.error(f"Failed to load {table_id}: {e}")
                # Continue loading other tables

        logger.info(f"Successfully loaded {len(results)}/{len(table_ids)} tables")
        return results

    def load_all_tables(self) -> Dict[str, Tuple[pd.DataFrame, Dict[str, str]]]:
        """
        Load all 16 Excel tables.

        Returns:
            Dictionary mapping table_id to (DataFrame, metadata) tuples
        """
        all_table_ids = [f"table{i:02d}" for i in range(1, 17)]
        return self.load_multiple_tables(all_table_ids)


# Convenience function for quick loading
def load_excel_table(
    table_id: str,
    data_dir: Optional[Path] = None
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """
    Convenience function to load a single Excel table.

    Args:
        table_id: Table identifier (e.g., "table01", "table08")
        data_dir: Optional data directory path

    Returns:
        Tuple of (DataFrame, metadata dictionary)
    """
    loader = ExcelLoader(data_dir=data_dir)
    return loader.load_excel_table(table_id)
